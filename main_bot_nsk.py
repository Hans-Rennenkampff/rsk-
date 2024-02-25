import threading
import eel
import pandas as pd
from openpyxl import load_workbook


def get_file():
    return r".\resourses\xl\raspisanie.xlsx"


@eel.expose
def group(group, day_week):
    """
    Функция вывода расписания на день одной группы
    :param sheet: лист расписания из файла Exel
    :param group: название группы из списка групп
    :param day_week: день недели
    :return: #arr массив расписания
    """
    file = get_file()
    wb = load_workbook(file)
    xl = pd.ExcelFile(file)
    sheet = wb[xl.sheet_names[0]]
    day_week = int(day_week)
    print(group, end=" ")
    for i in range(1, 250):
        for j in range(1, 250):
            if sheet.cell(row=i, column=j).value == group:
                column_group = j
    if day_week == 1:
        print("понедельник")
    elif day_week == 2:
        print("вторник")
    elif day_week == 3:
        print("среда")
    elif day_week == 4:
        print("четверг")
    elif day_week == 5:
        print("урааа пятница")
    else:
        print("суббота")
    day_week = day_week * 13 - 7
    arr = []
    a = []
    for i in range(day_week, day_week + 13):
        if sheet.cell(row=i, column=column_group + 1).value:
            a.append(sheet.cell(row=i, column=2).value)
            print(sheet.cell(row=i, column=2).value, end=" | ")  # урок
            a.append(sheet.cell(row=i, column=3).value)
            print(sheet.cell(row=i, column=3).value, end=" | ")  # время
            a.append(sheet.cell(row=i, column=column_group).value)
            print(sheet.cell(row=i, column=column_group).value, end=" | ")  # predmet
            a.append(sheet.cell(row=i, column=column_group + 1).value)
            print(sheet.cell(row=i, column=column_group + 1).value, end=" | ")  # teacher
            a.append(sheet.cell(row=i, column=column_group + 2).value)
            print(sheet.cell(row=i, column=column_group + 2).value)  # cab
            arr.append(a.copy())
            a.clear()
    eel.get_arr(arr)


def start_eel():
    eel.init('web')
    eel.start('index.html', mode=False)


if __name__ == '__main__':
    thread = threading.Thread(target=start_eel)
    thread.start()
    thread2 = threading.Thread(target=group, args=("group_name", "day_number"))
    thread2.start()