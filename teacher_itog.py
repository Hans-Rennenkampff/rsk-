import openpyxl
import pandas as pd
from openpyxl import load_workbook

def group1(group, day_week):
    """
    Функция вывода расписания на день одной группы
    :param sheet: лист расписания из файла Exel
    :param group: название группы из списка групп
    :param day_week: день недели
    :return: #arr массив расписания
    """

    file = r".\resourses\xl\raspisanie.xlsx"
    wb = load_workbook(file)
    xl = pd.ExcelFile(file)
    sheet = wb[xl.sheet_names[0]]
    day_week=int(day_week)
    for i in range(1, 250):
        for j in range(1, 250):
            if (sheet.cell(row=i, column=j).value == group):
                column_group = j

    day_week = day_week * 13 - 7
    arr=[]
    a=[]
    for i in range(day_week, day_week + 13):
            if (sheet.cell(row=i, column=column_group + 1).value):
                a.append(sheet.cell(row=i, column=2).value)
                a.append(sheet.cell(row=i, column=3).value)
                a.append(sheet.cell(row=i, column=column_group).value)
                a.append(sheet.cell(row=i, column=column_group + 1).value)
                a.append(sheet.cell(row=i, column=column_group + 2).value)
                arr.append(a.copy())
                a.clear()
    return arr

def search_teacher(file):
    """
    Функция поиска преподователей в файле Exel, расписание
    :return: #set выводит множество преподователей
    """
    # Открываем файл
    workbook = openpyxl.load_workbook(file)

    # Выбираем нужный лист (sheet)
    sheet = workbook["КОМ"]

    # Создаем пустое множество
    teachers = set()

    # Итерируемся по ячейкам всех строк, каждый третий столбец начиная с пятого
    for row in range(5, 84):
        for column in range(5, 54, 3):
            # Получаем значение ячейки
            cell_value = sheet.cell(row=row, column=column).value

            # Проверяем, что значение не пустое и не числовое
            if cell_value is not None and not isinstance(cell_value, (int, float)):
                # Если значение удовлетворяет условию, добавляем его в множество
                teachers.add(cell_value)
        for column in range(59, sheet.max_column, 3):
            # Получаем значение ячейки
            cell_value = sheet.cell(row=row, column=column).value

            # Проверяем, что значение не пустое и не числовое
            if cell_value is not None and not isinstance(cell_value, (int, float)):
                # Если значение удовлетворяет условию, добавляем его в множество
                teachers.add(cell_value)

    # Выводим результат на экран
    print(teachers)
    return teachers

def teachers_lessons(name_teachers, file):
    """
    Функция выводящая рассписание преподователя
    :param name_teachers: имя преподователя НСК
    :param file: excel расписание
    :return: array: строки с расписанием
    """

    workbook = openpyxl.load_workbook(file)
    sheet = workbook["КОМ"]

    arr = []#массив для вывода
    a = []#временный массив
    for row in range(5, 84):
        for column in range(5, 54, 3):
            cell_value = sheet.cell(row=row, column=column).value
            # Проверяем, что значение не пустое и не числовое
            if cell_value==name_teachers:
                if (row<(19)):
                    a.append("понедельник")
                elif (row<(31)):
                    a.append("вторник")
                elif (row < (45)):
                    a.append("среда")
                elif (row < (58)):
                    a.append("четверг")
                elif (row < (71)):
                    a.append("пятница")
                else:
                    a.append("суббота")
                a.append(sheet.cell(row=row, column=2).value)
                a.append(sheet.cell(row=row, column=3).value)
                a.append(sheet.cell(row=row, column=column - 1).value)
                a.append(sheet.cell(row=row, column=column).value)
                a.append(sheet.cell(row=row, column=column + 1).value)
                arr.append(a.copy())
                a.clear()
    for row in range(5, 84):
        for column in range(59, 193, 3):
            cell_value = sheet.cell(row=row, column=column).value
            # Проверяем, что значение не пустое и не числовое
            if cell_value==name_teachers:
                if (row<(19)):
                    a.append("понедельник")
                elif (row<(32)):
                    a.append("вторник")
                elif (row < (45)):
                    a.append("среда")
                elif (row < (58)):
                    a.append("четверг")
                elif (row < (71)):
                    a.append("пятница")
                else:
                    a.append("суббота")
                a.append(sheet.cell(row=row, column=2).value)
                a.append(sheet.cell(row=row, column=3).value)
                a.append(sheet.cell(row=row, column=column - 1).value)
                a.append(sheet.cell(row=row, column=column).value)
                if sheet.cell(row=row, column=column + 1).value:
                    a.append(sheet.cell(row=row, column=column + 1).value)
                else:
                    a.append("-")
                arr.append(a.copy())
                a.clear()
    print(arr)
    return arr

def bool_mat_f_graphics_t(name_teachers, file):
    """
    Boolean matrix for graphics массив булевых значений заполняющий графическую табличку расписания преподователя
    :param name_teachers: имя преподователя
    :param file: excel файл расписания
    :return: булевый массив 12*6
    """
    list = teachers_lessons(name_teachers, file)
    N, M = 6, 12
    day_week = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
    week=[]
    for i in range(N):
        week.append([])
    for i in range(len(list)):
        for j in range(N):
            if list[i][0] == day_week[j]:
                week[j].append(list[i][1])
    #print(week)
    timetable = [[False for _ in range(6)] for _ in range(12)]
    # Проходим по каждому уроку в каждый день и отмечаем его наличие в расписании
    for i, day in enumerate(week):
        for lesson in day:
            timetable[lesson-1][i] = True
    print(timetable)
    return timetable

def bool_mat_f_graphics(name_group, file):
    """

    :param name_group: название группы
    :param file: excel файл расписания
    :return: булевый массив 12*6
    """
    N, M = 6, 12
    day_week = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
    list = []
    for i in range(N):
        list.append(group1(name_group, i+1))
    week=[]
    for i in range(N):
        week.append([])
    for i in range(N):
        for j in range(len(list[i])):
            week[i].append(list[i][j][0])
    #print(week)
    timetable = [[False for _ in range(6)] for _ in range(12)]
    # Проходим по каждому уроку в каждый день и отмечаем его наличие в расписании
    for i, day in enumerate(week):
        for lesson in day:
            timetable[lesson-1][i] = True
    print(timetable)
    return timetable


file=r".\resourses\xl\raspisanie.xlsx"
print("Поиск всех преподователей в расписании НСК на неделю:")
search_teacher(r".\resourses\xl\raspisanie.xlsx")
print("Вывод расписания 1 преподователя:")
teachers_lessons("Ризванова А.Ф", file)
print("Вывод графической таблицы 1-го преподователя (Тагирова К.М, распечатанно внутри из-за функции):")
bool_mat_f_graphics_t("Тагиров К.М", file)
print("Вывод графической таблицы группы ИС-2:")
bool_mat_f_graphics("ИС-2", file)





