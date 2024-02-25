import eel
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import  os
import time
def countdown():
    while True:
        time.sleep(1)
        print("Осталось {} секунд".format(15 * 60 - time.time()))

def get_file():
    return r".\resourses\xl\raspisanie.xlsx"

@eel.expose
def group1(group, day_week):
    """
    Функция вывода расписания на день одной группы
    :param sheet: лист расписания из файла Exel
    :param group: название группы из списка групп
    :param day_week: день недели
    :return: #arr массив расписания
    """
    day_week = int(day_week)
    for i in range(1, 250):
        for j in range(1, 250):
            if (sheet.cell(row=i, column=j).value == group):
                column_group = j
    day_week = day_week * 13 - 7
    arr=[]
    a=[]
    for i in range(day_week, day_week + 13):
        if (sheet.cell(row=i, column=column_group + 1).value):
            a.append(sheet.cell(row=i, column=2).value)

            a.append(sheet.cell(row=i, column=3).value)

            a.append(sheet.cell(row=i, column=column_group).value)

            a.append(sheet.cell(row=i, column=column_group + 1).value)

            a.append(sheet.cell(row=i, column=column_group + 2).value)

            arr.append(a.copy())
            a.clear()
    return arr
@eel.expose
def bool_mat_f_graphics(name_group):
    """

    :param name_group: название группы
    :param file: excel файл расписания
    :return: булевый массив 12*6
    """
    N, M = 6, 12
    day_week = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
    list = []
    for i in range(N):
        list.append(group1(name_group, i+1))
    week=[]
    for i in range(N):
        week.append([])
    for i in range(N):
        for j in range(len(list[i])):
            week[i].append(list[i][j][0])
    #print(week)
    timetable = [[False for _ in range(6)] for _ in range(12)]
    # Проходим по каждому уроку в каждый день и отмечаем его наличие в расписании
    for i, day in enumerate(week):
        for lesson in day:
            timetable[lesson-1][i] = True
    print(timetable)
    return timetable
@eel.expose
def plot_bool_table(arr: np.ndarray, filename: str):
  """
  Функция генерации изображения расписания в виде таблицы
  :param arr: массив булевых значений в расписании, np.array
  :param filename: имя файла куда сохранять
  :return: сохраняет картинку заданной таблицы
  """
  # Устанавливаем размер фигуры
  plt.figure(figsize=(6, 12))

  # Закрашиваем клетки, где элементы массива равны True
  plt.imshow(arr, cmap='Reds', interpolation='nearest')

  # Добавляем сетку для ориентации
  plt.gca().set_xticks(np.arange(-0.5, arr.shape[1], 1), minor=True)
  plt.gca().set_yticks(np.arange(-0.5, arr.shape[0], 1), minor=True)
  plt.grid(which='minor', color='black', linestyle='-', linewidth=2)

  # Добавляем подписи на обеих осях
  plt.xticks(range(arr.shape[1]), ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"])
  plt.yticks(range(12), range(1, 13))

  # Сохраняем изображение
  plt.savefig(filename, bbox_inches='tight')

#как использовать
@eel.expose
def graph(group):
    name_file = group + ".png"
    output_dir = "web/output/"
    if not os.path.isfile(os.path.join(output_dir, name_file)):
        plot_bool_table(np.array(bool_mat_f_graphics(group)), output_dir + name_file)
        photo_file = open(output_dir + name_file, 'rb')
if __name__ == '__main__':
    file = get_file()
    wb = load_workbook(file)
    xl = pd.ExcelFile(file)
    sheet = wb[xl.sheet_names[0]]
    eel.init('web')
    eel.spawn(eel.start( 'index.html', mode=False, port=8080))
    countdown()